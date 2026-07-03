from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, auth as fb_auth, firestore
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os, io, tempfile
import urllib.request
import urllib.parse
import urllib.error
import json
import re
import hmac
import hashlib
import time
from PIL import Image as PILImage
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.middleware.proxy_fix import ProxyFix

# ── Firebase Admin SDK ────────────────────────────────────────
# Credenciais definidas como variável de ambiente no Railway
_firebase_initialized = False
_firebase_init_error  = None
_cred_json = os.environ.get('FIREBASE_ADMIN_CREDENTIALS')
if _cred_json:
    try:
        _cred = credentials.Certificate(json.loads(_cred_json))
        firebase_admin.initialize_app(_cred)
        _firebase_initialized = True
        print('Firebase Admin: inicializado com sucesso')
    except Exception as _e:
        _firebase_init_error = str(_e)
        print(f'Firebase Admin ERRO: {_e}')
else:
    _firebase_init_error = 'FIREBASE_ADMIN_CREDENTIALS nao definido'
    print('Firebase Admin: variavel FIREBASE_ADMIN_CREDENTIALS em falta')

# ── App & CORS ────────────────────────────────────────────────
# Aceita várias origens separadas por vírgula (GitHub Pages + Cloudflare Pages)
ALLOWED_ORIGINS = [o.strip() for o in os.environ.get(
    'ALLOWED_ORIGIN',
    'https://car-claim-ckd.pages.dev'
).split(',') if o.strip()]
CLOUDINARY_BASE = 'https://res.cloudinary.com/dos2jsgzg/'
TEMPLATE_PATH   = os.path.join(os.path.dirname(__file__), 'template.xlsx')

# Chave para o endpoint de exportação (Power BI / Power Apps).
EXPORT_API_KEY  = os.environ.get('EXPORT_API_KEY', '')

# Cloudinary — para assinar uploads (signed upload).
# Definir CLOUDINARY_API_KEY e CLOUDINARY_API_SECRET no painel do Render.
CLOUDINARY_CLOUD      = 'dos2jsgzg'
CLOUDINARY_API_KEY    = os.environ.get('CLOUDINARY_API_KEY', '')
CLOUDINARY_API_SECRET = os.environ.get('CLOUDINARY_API_SECRET', '')
CLOUDINARY_FOLDER     = 'garantia-car'

# ── Limites anti-DoS no processamento de imagem ───────────────
MAX_DOWNLOAD_BYTES = 15 * 1024 * 1024   # 15 MB por foto
PILImage.MAX_IMAGE_PIXELS = 40_000_000  # ~40 MP; acima disso o Pillow levanta erro (decompression bomb)

# Cliente Firestore (lazy) — só inicializa quando o endpoint é usado
_fs_client = None
def get_firestore():
    global _fs_client
    if _fs_client is None and _firebase_initialized:
        _fs_client = firestore.client()
    return _fs_client

app = Flask(__name__)
CORS(app,
     origins=ALLOWED_ORIGINS,
     methods=['GET', 'POST', 'OPTIONS'],
     allow_headers=['Content-Type', 'Authorization'])

# Render fica atrás de proxy → usa o IP real do cliente (X-Forwarded-For)
# para que o rate limiting seja por cliente, e não por IP do proxy.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)

# Rate limiting — anti abuso/brute force/DoS.
# Armazenamento em memória (adequado a 1 worker; usar Redis se escalar).
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=['300 per hour'],
    storage_uri='memory://',
)


# ── Headers de segurança em todas as respostas ────────────────
@app.after_request
def _security_headers(resp):
    resp.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    resp.headers['X-Content-Type-Options']    = 'nosniff'
    resp.headers['X-Frame-Options']           = 'DENY'
    resp.headers['Referrer-Policy']           = 'no-referrer'
    resp.headers['Content-Security-Policy']   = "default-src 'none'; frame-ancestors 'none'"
    return resp

# ── Auth ──────────────────────────────────────────────────────
def verify_token():
    """Verifica o Firebase ID token no header Authorization: Bearer <token>"""
    if not _firebase_initialized:
        print(f'verify_token: Firebase nao inicializado — {_firebase_init_error}')
        return None
    header = request.headers.get('Authorization', '')
    if not header.startswith('Bearer '):
        print('verify_token: header Authorization ausente ou mal formado')
        return None
    token = header.split('Bearer ', 1)[1].strip()
    try:
        # check_revoked: rejeita token de utilizador desativado/revogado
        return fb_auth.verify_id_token(token, check_revoked=True)
    except Exception as e:
        print(f'verify_token: token invalido — {e}')
        return None

# ── Validação ─────────────────────────────────────────────────
def is_valid_cloudinary_url(url):
    """Aceita apenas URLs do Cloudinary do projeto."""
    if not url or not isinstance(url, str):
        return False
    return url.startswith(CLOUDINARY_BASE)

def sanitize_str(value, max_len=500):
    if not value:
        return ''
    return str(value).strip()[:max_len]

def sanitize_int(value, default=1, min_val=0, max_val=9999):
    try:
        v = int(value if value is not None else default)
        return max(min_val, min(max_val, v))
    except (TypeError, ValueError):
        return default

def excel_safe(value):
    """Previne injeção de fórmula (CSV/Excel injection): se o texto começar
    com =, +, -, @ (ou tab/CR), o Excel/openpyxl tratariam como fórmula.
    Prefixa com ' para forçar interpretação como texto literal."""
    s = '' if value is None else str(value)
    if s and s[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + s
    return s

# ── Tradução ──────────────────────────────────────────────────
def translate_to_english(text):
    if not text:
        return ''
    text = text.strip()
    pt_words = ['de', 'do', 'da', 'em', 'no', 'na', 'ao', 'foi', 'com', 'para',
                'uma', 'um', 'que', 'por', 'nossa', 'nosso', 'este', 'esta',
                'detectamos', 'modelo', 'durante', 'processo', 'item', 'danos']
    is_portuguese = sum(1 for w in pt_words if f' {w} ' in f' {text.lower()} ') >= 2
    if not is_portuguese:
        return text.upper()
    try:
        params = urllib.parse.urlencode({
            'client': 'gtx', 'sl': 'pt', 'tl': 'en', 'dt': 't', 'q': text
        })
        url = f'https://translate.googleapis.com/translate_a/single?{params}'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            result = json.loads(response.read().decode('utf-8'))
            translated = ''.join([item[0] for item in result[0] if item[0]])
            return translated.upper()
    except Exception as e:
        print(f'Translation error: {e}')
        return text.upper()

# ── Processamento de imagem ───────────────────────────────────
def download_and_process(url, width, height):
    """Baixa imagem apenas de URLs Cloudinary e processa com Pillow."""
    if not is_valid_cloudinary_url(url):
        print(f'URL rejeitada (não Cloudinary): {url}')
        return None
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as response:
            clen = response.headers.get('Content-Length')
            if clen and int(clen) > MAX_DOWNLOAD_BYTES:
                print(f'Imagem rejeitada: Content-Length {clen} > limite')
                return None
            # Lê no máximo MAX_DOWNLOAD_BYTES+1 para detectar excesso mesmo sem Content-Length
            img_data = response.read(MAX_DOWNLOAD_BYTES + 1)
        if len(img_data) > MAX_DOWNLOAD_BYTES:
            print('Imagem rejeitada: excede limite de bytes')
            return None
        print(f'Downloaded {len(img_data)} bytes')
        img = PILImage.open(io.BytesIO(img_data))
        if img.mode in ('RGBA', 'P', 'LA'):
            img = img.convert('RGB')
        img.thumbnail((width, height), PILImage.LANCZOS)
        tmp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
        img.save(tmp.name, 'JPEG', quality=88)
        tmp.close()
        print(f'Processed image: {img.size} -> saved to {tmp.name}')
        return tmp.name
    except Exception as e:
        print(f'Image processing error: {e}')
        return None

# ── Rotas ─────────────────────────────────────────────────────
@app.route('/health', methods=['GET'])
@limiter.exempt
def health():
    # Resposta mínima — não expõe versões nem estado interno ao público
    if not _firebase_initialized:
        return jsonify({'status': 'degraded'}), 503
    return jsonify({'status': 'ok'})


@app.route('/sign-upload', methods=['POST'])
@limiter.limit('60 per hour')
def sign_upload():
    # Assina um upload Cloudinary. Só utilizadores autenticados (JWT Firebase)
    # obtêm assinatura — substitui o upload preset público (unsigned).
    user_token = verify_token()
    if not user_token:
        return jsonify({'error': 'Não autorizado. Faça login novamente.'}), 401
    if not CLOUDINARY_API_SECRET or not CLOUDINARY_API_KEY:
        return jsonify({'error': 'Cloudinary não configurado.'}), 503

    timestamp = int(time.time())
    # Restringe os formatos aceites no servidor (assinado) — impede subir
    # SVG/HTML/ficheiros arbitrários para a pasta. Inclui formatos do telemóvel.
    allowed_formats = 'jpg,jpeg,png,webp,gif,heic,heif'
    # Parâmetros a assinar, por ordem alfabética (regra do Cloudinary)
    to_sign   = (f'allowed_formats={allowed_formats}'
                 f'&folder={CLOUDINARY_FOLDER}&timestamp={timestamp}')
    signature = hashlib.sha1((to_sign + CLOUDINARY_API_SECRET).encode('utf-8')).hexdigest()

    return jsonify({
        'signature':      signature,
        'timestamp':      timestamp,
        'apiKey':         CLOUDINARY_API_KEY,
        'folder':         CLOUDINARY_FOLDER,
        'allowedFormats': allowed_formats,
        'cloudName':      CLOUDINARY_CLOUD,
    })


# ── Export para Power BI / Power Apps ─────────────────────────
def _check_export_key():
    """Valida a chave de API APENAS via header X-API-Key, de forma
    resistente a timing attacks. Devolve True se válida.
    A chave NÃO é aceite por query string (?key=) para não vazar em
    logs de servidor/proxy, histórico do navegador e header Referer."""
    if not EXPORT_API_KEY:
        return False  # sem chave configurada → endpoint fechado
    provided = request.headers.get('X-API-Key', '')
    return hmac.compare_digest(str(provided), EXPORT_API_KEY)


def _ms_to_iso(value):
    """Converte timestamp em milissegundos para ISO 8601 (ou None)."""
    if not value:
        return None
    try:
        return datetime.utcfromtimestamp(int(value) / 1000).isoformat() + 'Z'
    except (TypeError, ValueError, OverflowError):
        return None


@app.route('/export-data', methods=['GET'])
@limiter.limit('30 per hour')
def export_data():
    # 1. Autenticação por chave de API
    if not _check_export_key():
        return jsonify({'error': 'Não autorizado.'}), 401

    fs = get_firestore()
    if fs is None:
        return jsonify({'error': 'Serviço indisponível.'}), 503

    try:
        # 2. Lê todos os incidentes do Firestore
        docs = fs.collection('incidents').stream()
        rows = []
        for doc in docs:
            d = doc.to_dict() or {}
            photos = d.get('photos', []) or []
            history = d.get('history', []) or []

            rows.append({
                'id':            doc.id,
                'carNum':        d.get('carNum', ''),
                'status':        d.get('status', 'pending'),
                'incidentType':  d.get('incidentType', 'normal'),
                'partNo':        d.get('partNo', ''),
                'partName':      d.get('partName', ''),
                'model':         d.get('model', ''),
                'orderNo':       d.get('orderNo', ''),
                'lotNo':         d.get('lotNo', ''),
                'ngQty':         d.get('ngQty', ''),
                'defect':        d.get('defect', ''),
                'detected':      d.get('detected', ''),
                'user':          d.get('user', ''),
                'userId':        d.get('userId', ''),
                'tracking':      d.get('tracking', ''),
                'eta':           d.get('eta', ''),
                'chassisNote':   d.get('chassisNote', ''),
                'createdAt':     _ms_to_iso(d.get('createdAt')),
                'updatedAt':     _ms_to_iso(d.get('updatedAt')),
                'sentAt':        _ms_to_iso(d.get('sentAt')),
                'receivedAt':    _ms_to_iso(d.get('receivedAt')),
                'completedAt':   _ms_to_iso(d.get('completedAt')),
                'nPhotos':       len(photos),
                'photoUrls':     ' | '.join(p.get('url', '') for p in photos if isinstance(p, dict)),
                'history':       [
                    {
                        'status':    h.get('status', ''),
                        'note':      h.get('note', ''),
                        'user':      h.get('user', ''),
                        'isNote':    bool(h.get('isNote', False)),
                        'timestamp': _ms_to_iso(h.get('timestamp')),
                    }
                    for h in history if isinstance(h, dict)
                ],
            })

        return jsonify({
            'generatedAt': datetime.utcnow().isoformat() + 'Z',
            'count':       len(rows),
            'incidents':   rows,
        })

    except Exception as e:
        print(f'Error in export_data: {e}')
        return jsonify({'error': 'Erro ao exportar dados.'}), 500


@app.route('/export-partslist', methods=['GET'])
@limiter.limit('30 per hour')
def export_partslist():
    # Exporta a base de peças (partsDB) — pack lists importadas.
    # Mesma chave de API do /export-data.
    if not _check_export_key():
        return jsonify({'error': 'Não autorizado.'}), 401

    fs = get_firestore()
    if fs is None:
        return jsonify({'error': 'Serviço indisponível.'}), 503

    try:
        docs = fs.collection('partsDB').stream()
        rows = []
        for doc in docs:
            d = doc.to_dict() or {}
            rows.append({
                'id':         doc.id,
                'partNo':     d.get('partNo', ''),
                'partName':   d.get('partName', ''),
                'lotNo':      d.get('lotNo', ''),
                'orderNo':    d.get('orderNo', ''),
                'model':      d.get('model', ''),
                'qty':        d.get('qty', ''),
                'source':     d.get('source', ''),
                'importedAt': _ms_to_iso(d.get('importedAt')),
            })

        return jsonify({
            'generatedAt': datetime.utcnow().isoformat() + 'Z',
            'count':       len(rows),
            'parts':       rows,
        })

    except Exception as e:
        print(f'Error in export_partslist: {e}')
        return jsonify({'error': 'Erro ao exportar peças.'}), 500

# ── Admin: gestão de utilizadores (Firebase Admin SDK) ────────
VALID_ROLES = ('admin', 'user', 'viewer')


def verify_admin():
    """Devolve o token decodificado se o chamador for admin; senão None.
    O papel é lido do Firestore (users/{uid}.role) pelo Admin SDK."""
    tok = verify_token()
    if not tok:
        return None
    fs = get_firestore()
    if fs is None:
        return None
    try:
        snap = fs.collection('users').document(tok['uid']).get()
        if snap.exists and (snap.to_dict() or {}).get('role') == 'admin':
            return tok
    except Exception as e:
        print(f'verify_admin error: {e}')
    return None


@app.route('/admin/users', methods=['GET'])
def admin_list_users():
    if not verify_admin():
        return jsonify({'error': 'Apenas administradores.'}), 403
    fs = get_firestore()
    try:
        profiles = {d.id: (d.to_dict() or {}) for d in fs.collection('users').stream()}
        users = []
        for u in fb_auth.list_users().iterate_all():
            p = profiles.get(u.uid, {})
            users.append({
                'uid':       u.uid,
                'email':     u.email or p.get('email', ''),
                'name':      p.get('name') or (u.display_name or ''),
                'role':      p.get('role', 'user'),
                'disabled':  bool(u.disabled),
                'createdAt': p.get('createdAt'),
            })
        users.sort(key=lambda x: (x['name'] or x['email'] or '').lower())
        return jsonify({'users': users})
    except Exception as e:
        print(f'admin_list_users error: {e}')
        return jsonify({'error': 'Erro ao listar utilizadores.'}), 500


@app.route('/admin/users', methods=['POST'])
@limiter.limit('30 per hour')
def admin_create_user():
    tok = verify_admin()
    if not tok:
        return jsonify({'error': 'Apenas administradores.'}), 403
    data     = request.get_json(silent=True) or {}
    name     = sanitize_str(data.get('name', ''), 100)
    email    = sanitize_str(data.get('email', ''), 200).lower()
    password = str(data.get('password', ''))
    role     = data.get('role', 'user')
    if role not in VALID_ROLES:
        return jsonify({'error': 'Papel inválido.'}), 400
    if not email or '@' not in email:
        return jsonify({'error': 'Email inválido.'}), 400
    if len(password) < 8:
        return jsonify({'error': 'A senha deve ter pelo menos 8 caracteres.'}), 400
    try:
        u = fb_auth.create_user(email=email, password=password,
                                display_name=(name or None))
        get_firestore().collection('users').document(u.uid).set({
            'name': name, 'email': email, 'role': role,
            'createdAt': int(time.time() * 1000), 'createdBy': tok['uid'],
        })
        return jsonify({'uid': u.uid, 'email': email, 'name': name, 'role': role})
    except fb_auth.EmailAlreadyExistsError:
        return jsonify({'error': 'Este email já está registado.'}), 409
    except ValueError as e:
        return jsonify({'error': f'Dados inválidos: {e}'}), 400
    except Exception as e:
        print(f'admin_create_user error: {e}')
        return jsonify({'error': 'Erro ao criar utilizador.'}), 500


@app.route('/admin/users/role', methods=['POST'])
def admin_set_role():
    tok = verify_admin()
    if not tok:
        return jsonify({'error': 'Apenas administradores.'}), 403
    data = request.get_json(silent=True) or {}
    uid  = sanitize_str(data.get('uid', ''), 128)
    role = data.get('role', '')
    if role not in VALID_ROLES:
        return jsonify({'error': 'Papel inválido.'}), 400
    if not uid:
        return jsonify({'error': 'uid obrigatório.'}), 400
    if uid == tok['uid'] and role != 'admin':
        return jsonify({'error': 'Não pode remover o seu próprio acesso de admin.'}), 400
    try:
        get_firestore().collection('users').document(uid).set({'role': role}, merge=True)
        return jsonify({'uid': uid, 'role': role})
    except Exception as e:
        print(f'admin_set_role error: {e}')
        return jsonify({'error': 'Erro ao alterar papel.'}), 500


@app.route('/admin/users/disable', methods=['POST'])
def admin_disable_user():
    tok = verify_admin()
    if not tok:
        return jsonify({'error': 'Apenas administradores.'}), 403
    data     = request.get_json(silent=True) or {}
    uid      = sanitize_str(data.get('uid', ''), 128)
    disabled = bool(data.get('disabled', True))
    if not uid:
        return jsonify({'error': 'uid obrigatório.'}), 400
    if uid == tok['uid']:
        return jsonify({'error': 'Não pode desativar a sua própria conta.'}), 400
    try:
        fb_auth.update_user(uid, disabled=disabled)
        if disabled:
            # Revoga tokens → o acesso cai de imediato (com check_revoked)
            fb_auth.revoke_refresh_tokens(uid)
        return jsonify({'uid': uid, 'disabled': disabled})
    except Exception as e:
        print(f'admin_disable_user error: {e}')
        return jsonify({'error': 'Erro ao atualizar utilizador.'}), 500


@app.route('/admin/users/delete', methods=['POST'])
def admin_delete_user():
    tok = verify_admin()
    if not tok:
        return jsonify({'error': 'Apenas administradores.'}), 403
    data = request.get_json(silent=True) or {}
    uid  = sanitize_str(data.get('uid', ''), 128)
    if not uid:
        return jsonify({'error': 'uid obrigatório.'}), 400
    if uid == tok['uid']:
        return jsonify({'error': 'Não pode excluir a sua própria conta.'}), 400
    try:
        fb_auth.delete_user(uid)
        try:
            get_firestore().collection('users').document(uid).delete()
        except Exception:
            pass
        return jsonify({'uid': uid, 'deleted': True})
    except Exception as e:
        print(f'admin_delete_user error: {e}')
        return jsonify({'error': 'Erro ao excluir utilizador.'}), 500


@app.route('/generate-car', methods=['POST'])
@limiter.limit('60 per hour')
def generate_car():
    # 1. Autenticação obrigatória
    user_token = verify_token()
    if not user_token:
        return jsonify({'error': 'Não autorizado. Faça login novamente.'}), 401

    tmp_files = []
    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({'error': 'Dados inválidos.'}), 400

        # 2. Sanitização de todos os inputs
        car_num     = sanitize_str(data.get('carNum', '000/26'), 20)
        part_name   = sanitize_str(data.get('partName', ''), 200).upper()
        part_no     = sanitize_str(data.get('partNo', ''), 100).upper()
        model       = sanitize_str(data.get('model', ''), 100).upper()
        order_no    = sanitize_str(data.get('orderNo', ''), 100).upper()
        lot_no      = sanitize_str(data.get('lotNo', ''), 100).upper()
        ng_qty      = sanitize_int(data.get('ngQty', 1))
        defect      = sanitize_str(data.get('defect', ''), 500).upper()
        detected    = sanitize_str(data.get('detected', ''), 500)
        user        = sanitize_str(data.get('user', ''), 100).upper()
        replacement = sanitize_str(data.get('replacement', 'NEED'), 20)
        issue_date  = sanitize_str(data.get('issueDate', datetime.now().strftime('%d/%m/%Y')), 20)
        photos_raw  = data.get('photos', [])

        # 3. Validação das URLs das fotos (apenas Cloudinary)
        if not isinstance(photos_raw, list):
            return jsonify({'error': 'Formato de fotos inválido.'}), 400
        photos = [p for p in photos_raw[:10] if is_valid_cloudinary_url(p)]

        repl_qty    = 0 if replacement == 'NO NEED' else ng_qty
        detected_en = translate_to_english(detected)
        # Remove prefixos de label que a tradução pode gerar
        # ex: "HOW DETECTED:", "HOW IT WAS DETECTED:", "COMO FOI DETECTADO:"
        detected_en = re.sub(
            r'^(HOW\s+(IT\s+WAS\s+)?DETECTED\s*:\s*|COMO\s+FOI\s+DETECTAD[OA]\s*:\s*)',
            '', detected_en, flags=re.IGNORECASE
        ).strip()
        short_defect = part_name + (' (' + defect[:50] + ')' if defect else '')
        parts_desc = []
        if detected_en:
            parts_desc.append(detected_en + '.')
        parts_desc.append('PHOTOS ARE ATTACHED FOR YOUR REFERENCE.')
        full_desc = ' '.join(parts_desc)

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb['CAR']

        # Todas as células de texto passam por excel_safe (anti formula injection)
        ws['U2'] = excel_safe('IRU No. ' + car_num)
        ws['U4'] = excel_safe(user)
        ws['U5'] = excel_safe(issue_date)
        ws['G6'] = excel_safe(part_no)
        ws['U6'] = excel_safe(part_name)
        ws['AE4'] = excel_safe(part_name)
        ws['AE6'] = excel_safe(part_no)
        ws['G7'] = excel_safe(order_no)
        ws['U7'] = excel_safe(order_no)
        ws['D8'] = excel_safe(model)
        ws['K8'] = ng_qty
        ws['U8'] = excel_safe(lot_no)
        ws['G9'] = excel_safe(short_defect)
        ws['A11'] = excel_safe(full_desc)
        ws['V16'] = repl_qty

        photo_anchors = ['A16', 'Z22', 'A37', 'Z50', 'A63', 'Z63']

        for i, photo_url in enumerate(photos):
            anchor = photo_anchors[i] if i < len(photo_anchors) else f'A{16 + (i * 15)}'
            tmp_path = download_and_process(photo_url, 1400, 1050)
            if tmp_path:
                tmp_files.append(tmp_path)
                try:
                    xl_img = XLImage(tmp_path)
                    xl_img.anchor = anchor
                    ws.add_image(xl_img)
                    print(f'Photo {i+1} inserted at {anchor}')
                except Exception as e:
                    print(f'Photo {i+1} insert error: {e}')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        part_code = (part_no or 'PART').replace(' ', '_')[:20]
        car_code  = car_num.replace('/', '_')
        filename  = f'CAR_No_{car_code}_{part_code}.xlsx'

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        # 4. Sem stack trace em produção
        print(f'Error in generate_car: {e}')
        return jsonify({'error': 'Erro interno ao gerar o relatório.'}), 500

    finally:
        for f in tmp_files:
            try:
                os.unlink(f)
            except Exception:
                pass

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
